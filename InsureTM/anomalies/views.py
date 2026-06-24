import logging
import re
from pathlib import Path
from django.contrib.auth import get_user_model
from django.db.models import Q, TextField
from django.db.models.functions import Cast
from django.http import HttpResponse
from datetime import datetime
from fpdf import FPDF

from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.decorators import action

from notifications.models import Notification
from utils.email_service import send_anomaly_reported_email, send_anomaly_updated_email
from .models import Anomalie
from .serializers import AnomalieSerializer

logger = logging.getLogger(__name__)

PDF_FONT_CANDIDATES = (
    '/usr/share/fonts/truetype/freefont/FreeSans.ttf',
    '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
)
PDF_FONT_BOLD_CANDIDATES = (
    '/usr/share/fonts/truetype/freefont/FreeSansBold.ttf',
    '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
)

EXPORT_COLUMNS = (
    ('ID', 0.05),
    ('Titre', 0.22),
    ('Gravité', 0.08),
    ('Priorité', 0.07),
    ('Statut', 0.08),
    ('Test lié', 0.10),
    ('Release', 0.11),
    ('Campagne', 0.12),
    ('Créé par', 0.08),
    ('Date', 0.09),
)


def _resolve_pdf_font(candidates):
    for path in candidates:
        if Path(path).is_file():
            return path
    return None


class AnomalieViewSet(viewsets.ModelViewSet):
    queryset = Anomalie.objects.all()
    serializer_class = AnomalieSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

    def get_queryset(self):
        queryset = Anomalie.objects.select_related(
            'test_case',
            'test_case__campaign',
            'test_case__campaign__project',
            'cree_par',
        )
        
        if self.request.user.role == 'TESTER':
            queryset = queryset.filter(cree_par=self.request.user)

        search = self.request.query_params.get('search')
        impact = self.request.query_params.get('impact')
        campaign_id = self.request.query_params.get('campaign_id')

        if campaign_id:
            queryset = queryset.filter(test_case__campaign_id=campaign_id)

        if search:
            term = search.strip()
            if term.startswith('#'):
                term = term[1:].strip()

            if re.fullmatch(r'\d+', term):
                queryset = queryset.annotate(
                    pk_str=Cast('pk', TextField()),
                ).filter(pk_str__icontains=term)
            else:
                queryset = queryset.filter(
                    Q(titre__icontains=term) |
                    Q(description__icontains=term) |
                    Q(test_case__campaign__title__icontains=term) |
                    Q(test_case__test_case_ref__icontains=term) |
                    Q(cree_par__username__icontains=term) |
                    Q(cree_par__first_name__icontains=term) |
                    Q(cree_par__last_name__icontains=term)
                )

        if impact and impact != 'ALL' and impact != 'Tout':
            queryset = queryset.filter(impact=impact.upper())

        ordering = self.request.query_params.get('ordering')
        if ordering:
            queryset = queryset.order_by(ordering)

        return queryset

    def perform_create(self, serializer):
        instance = serializer.save(cree_par=self.request.user)

        test_case = instance.test_case
        if not (test_case and test_case.campaign):
            return

        campaign = test_case.campaign
        recipient = campaign.imported_by
        recipients = [recipient] if recipient else list(
            get_user_model().objects.filter(role='ADMIN')
        )

        for r in recipients:
            if r and r != self.request.user:
                Notification.objects.create(
                    recipient=r,
                    title="Nouvelle Anomalie",
                    message=f"{self.request.user.username} a signalé une anomalie sur {test_case.test_case_ref}",
                    type='anomaly_reported',
                    related_campaign=campaign,
                    related_object_id=instance.id,
                )
                if r.email:
                    send_anomaly_reported_email(r, self.request.user, instance, test_case)

    def perform_update(self, serializer):
        old_instance = self.get_object()
        old_status = old_instance.statut
        instance = serializer.save()
        
        # Notify stakeholders if status or impact changed
        if old_status != instance.statut or 'impact' in serializer.validated_data:
            test_case = instance.test_case
            if test_case and test_case.campaign:
                campaign = test_case.campaign
                manager = campaign.imported_by
                reporter = instance.cree_par
                
                recipients = set()
                if manager: recipients.add(manager)
                if reporter: recipients.add(reporter)
                recipients.discard(self.request.user)
                
                for r in recipients:
                    Notification.objects.create(
                        recipient=r,
                        title="Anomalie Mise à jour",
                        message=f"L'anomalie #{instance.id} a été mise à jour : {instance.statut}",
                        type='anomaly_reported',
                        related_campaign=campaign,
                        related_object_id=instance.id
                    )
                    if r.email:
                        send_anomaly_updated_email(r, self.request.user, instance)

    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        queryset = self.get_queryset()

        font_regular = _resolve_pdf_font(PDF_FONT_CANDIDATES)
        font_bold = _resolve_pdf_font(PDF_FONT_BOLD_CANDIDATES)
        font_family = 'FreeSans' if font_regular else 'helvetica'

        class PDF(FPDF):
            def header(self):
                self.set_fill_color(239, 68, 68)
                self.rect(0, 0, self.w, 5, 'F')
                self.set_font(font_family, 'B', 14)
                self.set_text_color(30, 41, 59)
                self.set_y(8)
                self.cell(0, 7, pdf_text("Rapport d'Anomalies - InsureTM"), ln=True, align='L')
                self.set_font(font_family, '', 8)
                self.set_text_color(100, 116, 139)
                generated_label = 'Généré le' if font_regular else 'Genere le'
                generated_at = datetime.now().strftime(
                    '%d/%m/%Y à %H:%M' if font_regular else '%d/%m/%Y a %H:%M'
                )
                self.cell(
                    0, 5,
                    pdf_text(f'{generated_label} {generated_at}'),
                    ln=True,
                    align='L',
                )
                self.ln(1)

            def draw_table_header(self, col_widths):
                self.set_font(font_family, 'B', 7)
                self.set_fill_color(248, 250, 252)
                self.set_text_color(71, 85, 105)
                x = self.l_margin
                y = self.get_y()
                for (header, _), width in zip(EXPORT_COLUMNS, col_widths):
                    self.set_xy(x, y)
                    self.cell(width, 7, pdf_text(header), border=1, align='L', fill=True)
                    x += width
                self.ln(7)

        def pdf_text(value):
            text = str(value or '-')
            if font_regular:
                return text
            return text.encode('latin-1', 'replace').decode('latin-1')

        pdf = PDF(orientation='L', unit='mm', format='A4')
        pdf.set_margins(4, 10, 4)
        pdf.set_auto_page_break(auto=True, margin=10)
        if font_regular:
            pdf.add_font(font_family, '', font_regular)
            pdf.add_font(font_family, 'B', font_bold or font_regular)
        pdf.add_page()

        available_w = pdf.w - pdf.l_margin - pdf.r_margin
        col_widths = [available_w * ratio for _, ratio in EXPORT_COLUMNS]

        pdf.draw_table_header(col_widths)

        row_h = 7
        fill = False
        for an in queryset:
            if pdf.get_y() > pdf.h - pdf.b_margin - row_h:
                pdf.add_page()
                pdf.draw_table_header(col_widths)

            values = self._anomaly_export_fields(an)
            x = pdf.l_margin
            y = pdf.get_y()
            for col_idx, ((_, _), width, value) in enumerate(zip(EXPORT_COLUMNS, col_widths, values)):
                pdf.set_xy(x, y)
                text = pdf_text(value)
                is_date_col = col_idx == len(EXPORT_COLUMNS) - 1
                if not is_date_col and col_idx != 0:
                    max_chars = max(6, int(width / 1.9))
                    if len(text) > max_chars:
                        text = text[: max_chars - 3] + '...'
                style = 'B' if col_idx == 0 else ''
                pdf.set_font(font_family, style, 7)
                if col_idx == 2:
                    if an.impact in ['CRITIQUE', 'BLOQUANTES']:
                        pdf.set_text_color(239, 68, 68)
                    elif an.impact in ['MAJEUR', 'MINEURS']:
                        pdf.set_text_color(234, 179, 8)
                    else:
                        pdf.set_text_color(59, 130, 246)
                else:
                    pdf.set_text_color(30, 41, 59)
                pdf.cell(width, row_h, text, border=1, align='L', fill=fill)
                x += width
            pdf.set_y(y + row_h)
            fill = not fill

        response = HttpResponse(bytes(pdf.output()), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="rapport_anomalies.pdf"'
        return response

    def _anomaly_export_fields(self, an):
        test_ref = an.test_case.test_case_ref if an.test_case else ''
        project_name = '-'
        campaign_title = '-'
        if an.test_case and an.test_case.campaign:
            campaign_title = an.test_case.campaign.title or '-'
            if an.test_case.campaign.project:
                project_name = an.test_case.campaign.project.name or '-'
        if an.cree_par:
            author = f"{an.cree_par.first_name} {an.cree_par.last_name}".strip() or an.cree_par.username
        else:
            author = 'Inconnu'
        return [
            an.id,
            str(an.titre or '').replace('\x00', ''),
            an.impact or 'A_DEFINIR',
            an.priorite or 'A_DEFINIR',
            an.statut or 'OUVERTE',
            test_ref or '-',
            project_name,
            campaign_title,
            author,
            an.cree_le.strftime('%d/%m/%Y'),
        ]

    @action(detail=False, methods=['get'])
    def export_xlsx(self, request):
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        queryset = self.get_queryset()
        headers = [
            'ID', 'Titre', 'Gravité', 'Priorité', 'Statut',
            'Test lié', 'Release', 'Campagne', 'Créé par', 'Date',
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = 'Anomalies'
        ws.append(headers)

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='1E293B')
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for an in queryset:
            ws.append(self._anomaly_export_fields(an))

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        min_widths = [8, 42, 14, 12, 14, 16, 22, 28, 18, 12]
        for col_idx, min_w in enumerate(min_widths, start=1):
            letter = get_column_letter(col_idx)
            max_len = min_w
            for cell in ws[letter]:
                if cell.value is not None:
                    max_len = max(max_len, min(len(str(cell.value)) + 2, 55 if col_idx == 2 else 35))
            ws.column_dimensions[letter].width = max_len

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="anomalies_export_{datetime.now().strftime("%Y-%m-%d")}.xlsx"'
        )
        return response

    @action(detail=False, methods=['post'])
    def diagnose_external_logs(self, request):
        logs = request.data.get('logs', '')
        code = request.data.get('code', '')
        
        if not logs:
            from rest_framework.response import Response
            return Response({'error': 'Les logs sont requis pour le diagnostic.'}, status=400)
            
        from analytics.groq_service import GroqService
        groq_service = GroqService()
        
        try:
            # S'il y a du code on l'intègre aux logs pour que l'IA comprenne le contexte
            full_context = logs
            if code:
                full_context = f"CODE TESTÉ:\n{code}\n\nLOGS D'ERREUR:\n{logs}"
                
            title, desc = groq_service.generate_anomaly_from_logs("Anomalie Manuelle / Externe", full_context)
            from rest_framework.response import Response
            return Response({
                "titre": title,
                "description": desc
            })
        except Exception as e:
            from rest_framework.response import Response
            return Response({"error": str(e)}, status=500)