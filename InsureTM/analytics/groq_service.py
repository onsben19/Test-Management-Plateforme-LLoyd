import os
import re
import groq
import base64
from django.db import connection
from campaigns.models import Campaign

class GroqService:
    def __init__(self):
        self.client = groq.Groq(
            api_key=os.environ.get("GROQ_API_KEY"),
        )
        
    # Groq model cascade — each model has its own daily quota
    # llama-3.3-70b-versatile : 1 000 req/day,  100 000 tokens/day  (best quality)
    # llama-3.1-8b-instant    : 14 400 req/day, 500 000 tokens/day  (fast, generous)
    # gemma2-9b-it            : 14 400 req/day, 500 000 tokens/day  (extra fallback)
    GROQ_FALLBACK_CHAIN = [
        "llama-3.3-70b-versatile",
        "llama-3.1-8b-instant",
        "openai/gpt-oss-20b",
        "openai/gpt-oss-120b",
    ]

    def _get_completion_with_fallback(self, messages, temperature=0.7, model_name="llama-3.3-70b-versatile"):
        # Build model chain: requested model first, then remaining fallbacks
        chain = [model_name] + [m for m in self.GROQ_FALLBACK_CHAIN if m != model_name]

        last_groq_error = None
        for model in chain:
            try:
                completion = self.client.chat.completions.create(
                    messages=messages,
                    model=model,
                    temperature=temperature,
                )
                return completion.choices[0].message.content.strip()
            except Exception as err:
                err_str = str(err)
                retryable = (
                    '429' in err_str
                    or '413' in err_str
                    or 'rate' in err_str.lower()
                    or 'quota' in err_str.lower()
                    or 'too large' in err_str.lower()
                    or 'decommissioned' in err_str.lower()
                    or 'model_not_found' in err_str.lower()
                )
                if retryable:
                    print(f"Groq model '{model}' unavailable ({err_str[:120]}), trying next model…")
                    last_groq_error = err
                    continue
                raise err

        # All Groq models exhausted → try Gemini cascade
        print(f"All Groq models exhausted. Falling back to Gemini… (last error: {last_groq_error})")
        gemini_api_key = os.environ.get("GEMINI_API_KEY")
        if not gemini_api_key:
            raise last_groq_error

        import google.generativeai as genai
        genai.configure(api_key=gemini_api_key)

        contents = []
        for msg in messages:
            if isinstance(msg['content'], list):
                for item in msg['content']:
                    if item['type'] == 'text':
                        contents.append(item['text'])
                    elif item['type'] == 'image_url':
                        b64 = item['image_url']['url'].split(',')[1]
                        image_bytes = base64.b64decode(b64)
                        contents.append({'mime_type': 'image/jpeg', 'data': image_bytes})
            else:
                contents.append(f"[{msg['role'].upper()}]: {msg['content']}")

        # Try multiple Gemini models in order (different quotas)
        last_gemini_err = None
        for gemini_model in ["gemini-1.5-flash", "gemini-2.0-flash", "gemini-2.5-flash"]:
            try:
                model = genai.GenerativeModel(gemini_model)
                response = model.generate_content(
                    contents,
                    generation_config=genai.types.GenerationConfig(temperature=temperature)
                )
                return response.text.strip()
            except Exception as gemini_err:
                last_gemini_err = gemini_err
                continue
        raise last_gemini_err
        
    def get_dynamic_schema(self, role, user_id):
        base_schema = """
        You are an expert PostgreSQL Data Analyst. Use the following database schema to answer user questions by generating a valid SQL query.
        
        Tables and Columns:
        """
        
        # 1. users_user - only visible to ADMIN
        if role == 'ADMIN':
            base_schema += "\n1. users_user (id, username, email, role, is_active, first_name, last_name)"
        
        base_schema += """
        2. campaigns_campaign (id, title, description, start_date, estimated_end_date, created_at, scheduled_at, nb_test_cases, project_id, imported_by_id)
        3. "testCases_testcase" (id, test_case_ref, data_json, status, campaign_id, tester_id, execution_date)
           - status values: 'PENDING', 'PASSED', 'FAILED'
           - data_json: JSON field containing extra details like 'Module', 'Domaine', 'Etape', 'Attendu' (e.g. data_json->>'Module' or inside array of objects).
        4. "Project_project" (id, name, description, status, created_at)
        5. anomalies_anomalie (id, titre, description, impact, priorite, visibilite, cree_le, test_case_id, cree_par_id)
           - impact values: 'BLOQUANTES', 'CRITIQUE', 'MAJEUR', 'MINEURS', 'COSMETIQUE', 'TEXTE', 'SIMPLE', 'FONCTIONNALITE'
        
        Rules:
        - Return ONLY the raw SQL query. NO markdown, NO code blocks, NO explanation.
        - Ensure table names are double-quoted if they contain mixed case or special characters.
        - CRITICAL: Always use double quotes around "testCases_testcase" and "Project_project" in your SQL queries. PostgreSQL is case-sensitive for table names, and omitting quotes will cause UndefinedTable errors.
        - CRITICAL: When filtering by string columns like 'title' or 'name', ALWAYS use ILIKE '%keyword%' instead of strict '=' equality to avoid case-sensitivity issues.
        - ALWAYS add a LIMIT 100 to any SELECT query to prevent overloading the database.
        - NEVER use DROP, DELETE, UPDATE, INSERT, or any data-modifying statements.
        """
        return base_schema

    def generate_sql(self, question, user):
        dynamic_schema = self.get_dynamic_schema(user.role, user.id)
        security_constraints = ""
        
        if user.role == 'ADMIN':
            security_constraints = "You are an ADMIN. Full access granted."
        elif user.role == 'MANAGER':
            security_constraints = f"You are a MANAGER (User ID: {user.id}). No access to users_user."
        else:
            security_constraints = f"You are a TESTER (User ID: {user.id}). Row-level security for your data."

        messages = [
            {"role": "system", "content": f"{dynamic_schema}\n\n{security_constraints}"},
            {"role": "user", "content": f"Generate a SQL query to answer: {question}"}
        ]
        sql_query = self._get_completion_with_fallback(messages, temperature=0)
        
        # Clean up markdown code block tags and surrounding conversation if present
        if "```" in sql_query:
            parts = sql_query.split("```")
            for part in parts:
                cleaned = part.strip()
                if cleaned.lower().startswith("sql"):
                    cleaned = cleaned[3:].strip()
                if cleaned and any(keyword in cleaned.upper() for keyword in ["SELECT", "WITH", "SHOW"]):
                    sql_query = cleaned
                    break
        else:
            sql_query = sql_query.strip()
            
        return sql_query

    def execute_query(self, sql_query):
        from datetime import datetime, date
        from decimal import Decimal

        def serialize_value(v):
            if isinstance(v, (datetime, date)): return v.isoformat()
            if isinstance(v, Decimal): return float(v)
            return v

        with connection.cursor() as cursor:
            cursor.execute(sql_query)
            if cursor.description is None:
                return []
            columns = [col[0] for col in cursor.description]
            results = cursor.fetchall()

        return [{k: serialize_value(v) for k, v in zip(columns, row)} for row in results]

    def build_simple_plotly_config(self, data, title=None):
        """Build a basic Plotly chart from SQL rows without calling an LLM."""
        if not data:
            return {"data": [], "layout": {"title": {"text": title or ""}}}

        keys = list(data[0].keys())
        label_key = keys[0]

        def is_numeric(value):
            if value is None or value == '':
                return False
            try:
                float(value)
                return True
            except (TypeError, ValueError):
                return False

        value_key = next(
            (k for k in keys[1:] if any(is_numeric(row.get(k)) for row in data)),
            label_key,
        )

        labels = [str(row.get(label_key, '')) for row in data]
        values = [float(row.get(value_key) or 0) for row in data]
        chart_type = "pie" if len(labels) <= 6 and all(v >= 0 for v in values) else "bar"

        trace = (
            {
                "type": "pie",
                "labels": labels,
                "values": values,
                "hole": 0.4,
                "textinfo": "percent+label",
            }
            if chart_type == "pie"
            else {
                "type": "bar",
                "x": labels,
                "y": values,
                "marker": {"color": "#3b82f6"},
            }
        )

        return {
            "data": [trace],
            "layout": {
                "title": {"text": title or ""},
                "paper_bgcolor": "transparent",
                "plot_bgcolor": "transparent",
                "margin": {"t": 40, "b": 80, "l": 40, "r": 20},
                "xaxis": {"tickangle": -30, "automargin": True},
                "autosize": True,
            },
        }

    def refresh_plotly_data(self, existing_config, raw_data, title=None):
        """Update an existing Plotly config with fresh SQL rows (no LLM)."""
        if not raw_data:
            if isinstance(existing_config, dict):
                return existing_config
            return {"data": [], "layout": {"title": {"text": title or ""}}}

        keys = list(raw_data[0].keys())
        label_key = keys[0]

        def is_numeric(value):
            if value is None or value == '':
                return False
            try:
                float(value)
                return True
            except (TypeError, ValueError):
                return False

        value_key = next(
            (k for k in keys[1:] if any(is_numeric(row.get(k)) for row in raw_data)),
            label_key,
        )
        labels = [str(row.get(label_key, '')) for row in raw_data]
        values = [float(row.get(value_key) or 0) for row in raw_data]

        if not isinstance(existing_config, dict) or not existing_config.get('data'):
            return self.build_simple_plotly_config(raw_data, title)

        layout = dict(existing_config.get('layout') or {})
        if title:
            layout['title'] = {'text': title}

        refreshed_traces = []
        for trace in existing_config.get('data', []):
            updated = dict(trace)
            trace_type = updated.get('type', 'bar')
            if trace_type == 'pie':
                updated['labels'] = labels
                updated['values'] = values
            elif trace_type in ('bar', 'scatter', 'line'):
                updated['x'] = labels
                updated['y'] = values
            refreshed_traces.append(updated)

        return {"data": refreshed_traces, "layout": layout}

    def _parse_document(self, uploaded_file):
        """Extract text content from common document formats."""
        try:
            filename = uploaded_file.name.lower()
            if filename.endswith('.pdf'):
                from pypdf import PdfReader
                reader = PdfReader(uploaded_file)
                text = ""
                for page in reader.pages:
                    text += page.extract_text() + "\n"
                return text
            
            elif filename.endswith('.docx'):
                import docx
                doc = docx.Document(uploaded_file)
                return "\n".join([p.text for p in doc.paragraphs])
            
            elif filename.endswith(('.xlsx', '.xls')):
                import openpyxl
                wb = openpyxl.load_workbook(uploaded_file)
                text = ""
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        text += "\t".join([str(cell) for cell in row if cell is not None]) + "\n"
                return text
            
            # Fallback for plain text
            return uploaded_file.read().decode('utf-8')
        except Exception as e:
            return f"[Error parsing document: {str(e)}]"

    def analyze_image(self, image_file, question):
        """Analyze an image using Groq's Vision model."""
        image_bytes = image_file.read()
        base64_image = base64.b64encode(image_bytes).decode('utf-8')
        
        messages = [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": f"En tant qu'expert QA, analyse cette capture d'écran et réponds à : {question}"},
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}
                    }
                ]
            }
        ]
        return self._get_completion_with_fallback(messages, temperature=0.7, model_name="llama-3.2-11b-vision-preview")

    def generate_plotly_config(self, data, question):
        """Generate a Plotly.js configuration JSON based on data and question."""
        prompt = f"""
        Given the following data: {data}
        And the user's question: {question}
        
        Generate a high-quality Plotly.js configuration in JSON format.
        Return ONLY the JSON object with 'data' and 'layout' keys.
        
        CRITICAL SPATIAL RULES (Avoid 'Spacing Error'):
        1. TITLE: MANDATORY: Set 'title': {{ 'text': question }} in layout. It will be rendered as a header.
        2. LEGEND: ALWAYS use 'legend': {{ 'orientation': 'h', 'y': -0.2, 'x': 0.5, 'xanchor': 'center', 'font': {{ 'size': 13 }} }}.
        3. MARGINS: Use 'margin': {{ 't': 10, 'b': 20, 'l': 10, 'r': 10 }} for maximum chart diameter.
        4. AXIS: For Bar/Line, use 'xaxis': {{ 'tickangle': -30, 'automargin': true, 'tickfont': {{ 'size': 12 }} }}.
        5. PIE/DONUT: MANDATORY for large visibility: 
           - Set 'hole': 0.4.
           - Set 'textinfo': 'percent+label'.
           - Set 'textposition': 'outside'.
           - For 3 or fewer slices, use 'pull': [0.03, 0.03, 0.03] to make it prominent.
        6. RADAR/POLAR: In layout, use 'polar': {{ 'radialaxis': {{ 'visible': true }}, 'angularaxis': {{ 'tickfont': {{ 'size': 14 }} }} }}.
        7. COLORS: Use the vibrant palette: ['#3b82f6', '#ec4899', '#8b5cf6', '#10b981', '#f59e0b', '#06b6d4', '#f43f5e'].
        8. FONT: Set global font color to '#f1f5f9' (slate-100).
        9. BACKGROUND: Use 'paper_bgcolor': 'transparent', 'plot_bgcolor': 'transparent'.
        10. AUTO-SCALING: 'autosize': true.
        11. THEME: Create a compact and readable chart. Ensure it fits well within a standard chat message container.
        12. SPACING: For charts with 2+ series, use 'barmode': 'group'. Ensure labels don't overlap. Adjust margins so no text is cut off.
        
        Example Output: {{"data": [...], "layout": {{ ... }} }}
        """
        
        messages = [{"role": "user", "content": prompt}]
        config_text = self._get_completion_with_fallback(messages, temperature=0)
        
        import re
        match = re.search(r'\{.*\}', config_text, re.DOTALL)
        if match:
            config_text = match.group(0)
            
        return config_text

    def process_query(self, question, user, uploaded_file=None, history=None):
        try:
            # Case 1: Vision analysis or document parsing if file is provided
            if uploaded_file:
                is_image = uploaded_file.name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp'))
                if is_image:
                    answer = self.analyze_image(uploaded_file, question)
                    return {"answer": answer, "type": "text", "sql": "", "data": []}
                else:
                    # Parse document and use Llama-3 to analyze content
                    doc_content = self._parse_document(uploaded_file)
                    prompt = f"""
                    Tu es un expert QA. On t'a fourni le document suivant : {uploaded_file.name}
                    
                    CONTENU DU DOCUMENT :
                    ---
                    {doc_content[:10000]} # Limit to 10k chars to avoid token issues
                    ---
                    
                    Question de l'utilisateur : {question}
                    
                    Consigne : Analyse le document par rapport à la question et réponds de manière précise. Si le document ne contient pas l'info, mentionne-le.
                    """
                    messages = [{"role": "user", "content": prompt}]
                    answer = self._get_completion_with_fallback(messages, temperature=0.7)
                    return {
                        "answer": answer,
                        "type": "text", "sql": "", "data": []
                    }
            
            # Intent Classification via LLM
            intent = "TEXT"
            intent_prompt = f"""
            You are an advanced classification model for a QA Platform Assistant.
            Analyze the user query and classify its intent.
            
            Query: "{question}"
            
            Available intents:
            - READINESS: The user is asking about readiness score, release readiness, deployment status, or estimated launch confidence.
            - SQL: The user is asking for database statistics, counts of campaigns, lists of projects, failures, anomalies, testers, execution dates, or any data stored in the database.
            - TEXT: The user is greeting you, asking a general QA question, asking you to write/reformulate an email, or engaging in general text conversation.
            
            Respond with ONLY one word from: READINESS, SQL, TEXT. Do not write anything else.
            """
            try:
                intent_res = self._get_completion_with_fallback([{"role": "user", "content": intent_prompt}], temperature=0)
                intent_res_cleaned = intent_res.strip().upper()
                if "READINESS" in intent_res_cleaned:
                    intent = "READINESS"
                elif "SQL" in intent_res_cleaned:
                    intent = "SQL"
                else:
                    intent = "TEXT"
            except Exception as classify_error:
                print(f"Classification failed: {classify_error}. Falling back to keywords.")
                # Fallback keyword matching
                readiness_keywords = ['score', 'readiness', 'prêt', 'déploiement', 'confiance', 'readynace']
                if any(kw in question.lower() for kw in readiness_keywords):
                    intent = "READINESS"
                else:
                    sql_keywords = ['combien', 'liste', 'nombre', 'qui', 'projet', 'campagne', 'anomalie', 'test', 'tester', 'moyen', 'taux']
                    if any(kw in question.lower() for kw in sql_keywords):
                        intent = "SQL"
                    else:
                        intent = "TEXT"

            # Route by Intent
            # Case 2: Readiness Score Intent
            if intent == "READINESS":
                campaign = Campaign.objects.order_by('-created_at').first()
                if campaign:
                    from analytics.readiness_service import ReleaseReadinessManager
                    readiness = ReleaseReadinessManager().calculate_readiness_score(campaign.id)
                    
                    prompt = f"""
                    Expert QA Platform Analyser. 
                    Explique le 'Release Readiness Score' au manager.
                    
                    Campagne: {campaign.title}
                    Score Actuel: {readiness.get('score', 0)}%
                    Répartition: {readiness.get('breakdown', {})}
                    Raisons précises: {readiness.get('reasons', [])}
                    
                    Question de l'utilisateur: {question}
                    
                    Règles:
                    1. Réponds en Français de manière très objective, professionnelle et concise.
                    2. Utilise les 'Raisons précises' pour justifier pourquoi le score est à ce niveau.
                    3. À la fin, propose TOUJOURS une action concrète (ex: reformuler une notification, contacter un testeur, etc.).
                    4. Si le score est < 80%, sois vigilant sur les risques.
                    5. NE dis JAMAIS "En tant qu'expert...". Va droit au but.
                    """
                    messages = [{"role": "user", "content": prompt}]
                    answer = self._get_completion_with_fallback(messages, temperature=0.7)
                    return {
                        "answer": answer,
                        "type": "text", 
                        "sql": "N/A (Calcul de score interne)", 
                        "data": readiness
                    }
                else:
                    return {
                        "answer": "Aucune campagne n'a été trouvée dans le système pour calculer le score de readiness.",
                        "type": "text",
                        "sql": "N/A",
                        "data": []
                    }

            # Case 3: Advanced Data query with SQL generation & fallback
            elif intent == "SQL":
                sql_query = self.generate_sql(question, user)
                try:
                    data = self.execute_query(sql_query)
                except Exception as sql_error:
                    print(f"SQL execution error: {sql_error}. Query was: {sql_query}")
                    # Friendly technical explanation fallback
                    fallback_prompt = f"""
                    L'utilisateur a posé une question sur les données : "{question}".
                    La requête SQL générée `{sql_query}` a échoué avec l'erreur : {str(sql_error)}.
                    
                    Réponds poliment en français, de manière très objective et professionnelle. Explique qu'il y a eu un problème technique lors de la récupération des données réelles, et invite l'utilisateur à reformuler ou à ajuster directement la requête SQL via le panneau de modifications. NE dis PAS "En tant que...".
                    """
                    answer = self._get_completion_with_fallback([{"role": "user", "content": fallback_prompt}], temperature=0.7)
                    return {
                        "answer": answer,
                        "type": "text",
                        "sql": sql_query,
                        "data": []
                    }

                if len(data) > 0:
                    # Secondary Cognitive Step: Interpret results
                    interpretation_prompt = f"""
                    Analyse ces résultats de données extraits de la plateforme pour répondre à la question de l'utilisateur.
                    
                    QUESTION : {question}
                    DONNÉES BRUTES : {data[:20]} (échantillon)
                    
                    CONSIGNES STRICTES :
                    1. Ne te contente pas de lister les chiffres. Identifie une CORRÉLATION, une TENDANCE ou un RISQUE caché.
                    2. Adopte un ton direct, objectif et très professionnel.
                    3. Propose une ACTION immédiate basée sur cette analyse technique.
                    4. Réponds en 2-3 phrases percutantes en Français.
                    5. NE MENTIONNE JAMAIS explicitement ton rôle. Ne commence JAMAIS ta réponse par "En tant que...". Réponds directement.
                    """
                    messages = [{"role": "user", "content": interpretation_prompt}]
                    interpretation = self._get_completion_with_fallback(messages, temperature=0.5)
                    expert_answer = f"Voici l'analyse des données :\n\n{interpretation}"
                else:
                    expert_answer = "Aucune donnée n'a été trouvée pour cette période ou ces critères. Un audit plus large est conseillé."

                # Decide if we use Plotly for complex requests
                complex_keywords = ['radar', 'scatter', 'bulles', 'comparative', 'avancé', 'plotly', 'graphique', 'chart', 'graphe', 'graphical']
                is_complex = any(kw in question.lower() for kw in complex_keywords)
                
                if is_complex and len(data) > 0:
                    plotly_config = self.generate_plotly_config(data, expert_answer) # Pass expert answer as context
                    return {
                        "answer": expert_answer,
                        "sql": sql_query,
                        "data": plotly_config,
                        "type": "plotly"
                    }
                
                # Default heuristics for simple charts
                chart_type = "table"
                if len(data) > 0:
                    if len(data) == 1 and len(data[0]) == 1: chart_type = "metric"
                    elif any(k in str(data[0].keys()).lower() for k in ["count", "total", "nb"]): chart_type = "bar"
                    elif any(k in str(data[0].keys()).lower() for k in ["date", "time", "day", "mois"]): chart_type = "line"
                
                return {
                    "answer": expert_answer,
                    "sql": sql_query,
                    "data": data,
                    "type": chart_type
                }

            # Case 4: General Chat Intent (TEXT)
            else:
                if history:
                    messages = list(history)
                    has_system = any(msg.get('role') == 'system' for msg in messages)
                    if not has_system:
                        messages.insert(0, {
                            "role": "system",
                            "content": """Tu es un assistant IA général et polyvalent.
                            Ton objectif est de répondre à n'importe quelle question posée par l'utilisateur.
                            1. Formatage Strict : Utilise le Markdown pour structurer tes réponses (gras, listes à puces, code).
                            2. Concision et Clarté : Sois direct, clair et pertinent.
                            3. Polyvalence Totale : Tu réponds de manière générale, sans être limité à une base de données. Tu ne génères pas de graphiques.
                            4. Langue : Tu t'exprimes en Français par défaut, de manière chaleureuse et professionnelle."""
                        })
                    # Add user's latest query if not already there
                    if not messages or messages[-1].get('role') != 'user' or messages[-1].get('content') != question:
                        messages.append({"role": "user", "content": question})
                else:
                    system_prompt = """Tu es un assistant IA général et polyvalent chez Lloyd Assurances.
                    Réponds de manière professionnelle, chaleureuse et concise aux questions de l'utilisateur.
                    Exprime-toi en Français par défaut."""
                    messages = [
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": question}
                    ]
                
                answer = self._get_completion_with_fallback(messages, temperature=0.7)
                return {
                    "answer": answer,
                    "type": "text",
                    "sql": "",
                    "data": []
                }

        except Exception as e:
            return {
                "answer": f"Erreur d'analyse : {str(e)}",
                "type": "error", "sql": "", "data": []
            }

    def _clean_email_reformulation(self, original, result, is_subject=False):
        """Cleanup for email reformulation — preserve facts, strip LLM artifacts."""
        result = result.strip().strip('"\'')
        result = re.sub(r'^(?:Objet|Subject)\s*:\s*', '', result, flags=re.I)
        result = re.sub(r'^(?:Voici (?:le|la)|Message reformulé\s*:)\s*', '', result, flags=re.I)
        result = re.sub(r'\[(?:Votre nom|Nom|Prénom|Date|Prenom)\]', '', result, flags=re.I)

        if is_subject:
            result = result.replace('**', '').replace('\n', ' ').replace('\r', ' ').strip()
            result = re.sub(r'\s{2,}', ' ', result)
            return result.strip(' .')

        email_fluff = (
            r'\s*(?:Je reste à votre entière disposition|Je reste à votre disposition)\s*\.?\s*$',
            r'\s*N\'hésitez pas(?: à me contacter)?(?: si besoin)?\s*\.?\s*$',
        )
        for pattern in email_fluff:
            if not re.search(pattern, original, flags=re.I):
                result = re.sub(pattern, '', result, flags=re.I).strip()

        if len(result) > len(original) * 1.7 + 40:
            return original.strip()

        return result.strip()

    def _clean_chat_reformulation(self, original, result):
        """Light cleanup for chat — stay close to what the user wrote."""
        result = result.strip().strip('"\'')
        result = re.sub(r'\[(?:Votre nom|Nom|Prénom|Date|Prenom)\]', '', result, flags=re.I)

        # Remove greetings/closings added by the model if absent in original
        if not re.match(r'^(?:Bonjour|Salut|Hello|Bonsoir)\b', original, flags=re.I):
            result = re.sub(r'^(?:Bonjour|Salut|Hello|Bonsoir)\s*,?\s*', '', result, flags=re.I).strip()
        if not re.search(r'(?:Cordialement|Merci(?: beaucoup)?)\s*\.?\s*$', original, flags=re.I):
            result = re.sub(
                r'\s*(?:Cordialement|Bien cordialement|Merci beaucoup)\s*\.?\s*$',
                '',
                result,
                flags=re.I,
            ).strip()

        email_fluff = (
            r'\s*(?:Je reste à votre disposition|N\'hésitez pas(?: à me contacter)?)\s*\.?\s*$',
            r'\s*Merci de votre attention\s*\.?\s*$',
        )
        for pattern in email_fluff:
            if not re.search(pattern, original, flags=re.I):
                result = re.sub(pattern, '', result, flags=re.I).strip()

        if len(result) > len(original) * 1.4 + 15:
            return original.strip()

        return result.strip()

    def reformulate_message(self, text, is_subject=False, is_test_steps=False, is_chat=False, is_email=False):
        if is_test_steps:
            system_prompt = (
                "Tu es un expert QA (Quality Assurance). Reformule les étapes de test suivantes en respectant EXACTEMENT ce format strict (en français) :\n\n"
                "Objectif : [Résumé clair et concis de ce qu'on teste en une phrase]\n\n"
                "Étapes :\n"
                "1. [Action détaillée (ex: Ouvrir l'URL : https://...)]\n"
                "2. [Action détaillée (ex: Saisir la valeur...)]\n"
                "3. [Vérification (ex: Vérifier que...)]\n\n"
                "Ne retourne QUE ce texte avec 'Objectif :' et 'Étapes :', rien d'autre. Pas de blabla d'introduction."
            )
        else:
            if is_subject:
                system_prompt = (
                    "Tu reformules un OBJET D'EMAIL professionnel en français (contexte QA / tests logiciels). "
                    "Une seule ligne, clair et concis. "
                    "Garde EXACTEMENT les identifiants techniques (ex: TC-AUTH-001, ANO-42). "
                    "Corrige l'orthographe sans changer le sens. "
                    "Pas de guillemets. Pas de markdown. Pas de retour à la ligne. "
                    "Pas de préfixe \"Objet:\".\n\n"
                    "Exemples :\n"
                    "- \"echec test connexion\" → \"Échec du test de connexion\"\n"
                    "- \"tc-auth-001 failed\" → \"Échec — TC-AUTH-001\"\n"
                    "- \"anomalie bloquante release 2\" → \"Anomalie bloquante — Release 2\""
                )
            else:
                if is_chat:
                    system_prompt = (
                        "Tu fais une reformulation LÉGÈRE d'un message de chat entre collègues. "
                        "Corrige l'orthographe, les accents et la grammaire. "
                        "Fluidifie la phrase seulement si nécessaire. "
                        "Garde le MÊME sens, le MÊME ton et le MÊME niveau de langue que l'auteur. "
                        "Si l'auteur tutoie (tu/te/ton), conserve le tutoiement. "
                        "Si l'auteur vouvoie, conserve le vouvoiement. "
                        "Ne réécris pas le message de zéro. Ne le rends pas plus sec ni plus direct. "
                        "N'ajoute pas de Bonjour, Cordialement ou formule de politesse si elles ne sont pas déjà dans le texte. "
                        "Ne change pas les identifiants (TC-AUTH-001, etc.). Pas de placeholder. Pas de markdown. Pas de guillemets.\n\n"
                        "Exemples :\n"
                        "- \"jai fini le test ca marche pas\" → \"J'ai fini le test, ça ne marche pas.\"\n"
                        "- \"tu peux regarder stp?\" → \"Tu peux regarder, stp ?\"\n"
                        "- \"salut est ce que tu as vu lanomalie\" → \"Salut, est-ce que tu as vu l'anomalie ?\"\n"
                        "- \"le cas tc-auth-001 a echoue\" → \"Le cas TC-AUTH-001 a échoué.\"\n\n"
                        "Retourne UNIQUEMENT le message reformulé."
                    )
                elif is_email:
                    system_prompt = (
                        "Tu reformules le CORPS d'un email professionnel en français (contexte QA / assurance / tests logiciels). "
                        "Corrige l'orthographe et la grammaire, améliore la clarté, ton professionnel sobre. "
                        "Garde EXACTEMENT le même sens et toutes les informations factuelles : "
                        "IDs de cas de test, noms de campagnes, statuts (PASSED/FAILED), numéros d'anomalie, dates, noms propres. "
                        "Ne modifie pas les identifiants techniques (TC-AUTH-001 reste TC-AUTH-001). "
                        "Structure simple : salutation courte si absente (Bonjour,), corps en 1 à 3 phrases, formule courte (Cordialement ou Merci). "
                        "Ne pas allonger inutilement. Ne pas inventer de détails. Pas de placeholder [Votre nom]. Pas de markdown.\n\n"
                        "Exemples :\n"
                        "- \"bonjour le test tc-auth-001 a echoue\" → \"Bonjour,\\n\\nLe cas de test TC-AUTH-001 a échoué.\\n\\nCordialement\"\n"
                        "- \"pouvez vous verifier lanomalie 42 stp\" → \"Bonjour,\\n\\nPouvez-vous vérifier l'anomalie 42, s'il vous plaît ?\\n\\nCordialement\"\n"
                        "- \"merci de traiter le bug login avant vendredi\" → \"Bonjour,\\n\\nMerci de traiter le bug de connexion avant vendredi.\\n\\nCordialement\"\n\n"
                        "Retourne UNIQUEMENT le corps de l'email reformulé."
                    )
                else:
                    system_prompt = (
                        "Tu es un assistant de rédaction professionnelle en français. "
                        "Reformule le texte suivant de façon SIMPLE et NATURELLE : corrige l'orthographe et la grammaire, "
                        "améliore légèrement la clarté, garde le même sens et un ton professionnel mais sobre. "
                        "RÈGLES STRICTES : "
                        "Ne pas allonger inutilement le message. "
                        "Ne pas ajouter de détails, d'arguments ou de phrases que l'auteur n'a pas mentionnés. "
                        "Ne pas ajouter de placeholder du type [Votre nom], [Nom], [Date]. "
                        "Structure légère : salutation courte si absente, corps du message, formule de politesse courte. "
                        "Pas de markdown. Pas de titre. Pas de commentaire hors message. "
                        "Retourne UNIQUEMENT le texte reformulé."
                    )
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": text}
        ]
        temperature = 0.1 if is_chat else (0.15 if is_email else 0.2)
        result = self._get_completion_with_fallback(messages, temperature=temperature, model_name="llama-3.1-8b-instant")
        if is_subject:
            result = self._clean_email_reformulation(text, result, is_subject=True)
        elif is_chat:
            result = self._clean_chat_reformulation(text, result)
        elif is_email:
            result = self._clean_email_reformulation(text, result, is_subject=False)
        else:
            result = re.sub(r'\[Votre nom\]', '', result, flags=re.I)
            result = re.sub(r'\[Nom\]', '', result, flags=re.I)
            result = result.strip()
        return result
    def generate_dashboard_brief(self, stats):
        """Generate a contextual AI brief for the manager dashboard with 5-min time-based rotation and deep data."""
        import time
        
        themes = [
            {"name": "Gestion des Risques", "focus": "les anomalies critiques, les échecs de tests et les retards potentiels.", "target_id": "recent-activity"},
            {"name": "Performance & Vélocité", "focus": "le volume d'exécutions, le taux de succès et la rapidité d'avancement des campagnes.", "target_id": "weekly-activity"},
            {"name": "Qualité logicielle", "focus": "la répartition des anomalies, la robustesse des tests passés et la couverture critique.", "target_id": "manager-anomaly-dist"},
            {"name": "Vision Stratégique", "focus": "l'équilibre entre volume d'exécution, bugs ouverts et état de préparation global (Readiness).", "target_id": "ml-timeline-guard"},
        ]
        
        # Consistent rotation every 5 minutes (300 seconds)
        theme_index = int(time.time() / 300) % len(themes)
        theme = themes[theme_index]
        
        # Format stats for prompt
        active_projects = stats.get('active_projects', 0)
        total_campaigns = stats.get('total_campaigns', 0)
        open_anomalies = stats.get('open_anomalies', 0)
        critical_anomalies = stats.get('critical_impact_count', 0)
        total_passed = stats.get('total_passed', 0)
        total_failed = stats.get('total_failed', 0)
        total_executions = stats.get('total_executions', 0)
        success_rate = stats.get('success_rate', 0)
        readiness_score = stats.get('readiness_score', 0)
        
        prompt = f"""
        En tant qu'expert QA Platform Analyser, génère un brief exécutif complet pour un Manager.
        
        ANGLE D'ANALYSE PRIORITAIRE : {theme['name']} (Ton analyse doit se concentrer particulièrement sur {theme['focus']})
        
        DONNÉES TEMPS RÉEL DU DASHBOARD :
        - État global : {total_executions} exécutions ({total_passed} passés, {total_failed} échoués).
        - Santé : {open_anomalies} anomalies ouvertes (DONT {critical_anomalies} à impact Critique/Bloquant).
        - Volume : {total_campaigns} campagnes sur {active_projects} projets.
        - Performance : {success_rate}% de succès / Readiness Score : {readiness_score}%.
        
        Règles :
        1. Réponds en Français de manière extrêmement professionnelle, concise et percutante (max 3-4 courtes phrases).
        2. Adopte la perspective du thème : {theme['name']}.
        3. FAIS LE LIEN entre les exécutions (succès/échecs) et les anomalies (surtout les critiques).
        4. Identifie immédiatement si la situation globale est 'Critique', 'Stable' ou 'Optimale' sous cet angle.
        5. Donne une recommandation stratégique directe.
        
        Format de réponse attendu : "Analyse {theme['name']} : [Statut]. [Analyse transversale liant exécutions et anomalies]. [Action prioritaire]."
        """
        try:
            messages = [{"role": "user", "content": prompt}]
            # Use fast/cheap model for brief — short output, doesn't need 70b quality
            brief_text = self._get_completion_with_fallback(messages, temperature=0.7, model_name="llama-3.1-8b-instant")
            return {
                "brief": brief_text,
                "target_id": theme['target_id'],
                "theme_name": theme['name']
            }
        except Exception as e:
            return {
                "brief": f"Analyse indisponible : {str(e)}",
                "target_id": None,
                "theme_name": "Inconnu"
            }

    def generate_playwright_test(self, test_title, test_data_json):
        """
        Génère un script de test Playwright à partir de données JSON extraites d'Excel.
        """
        prompt = f"""
        Tu es un expert QA en Automatisation avec Playwright (TypeScript).
        Ton objectif est de générer un script de test exécutable parfait à partir de données brutes.
        
        Titre du Cas de Test : {test_title}
        
        DONNÉES DU TEST (Étapes ou JSON) :
        {test_data_json}
        
        CONSIGNES STRICTES :
        1. ÉTAPE COGNITIVE DE PRÉ-CONSTRUCTION : Analyse et reformule intérieurement les étapes de test brutes ou désordonnées en une séquence d'actions et de vérifications logique et propre.
        2. Construit une séquence d'actions Playwright logique basée sur cette reformulation.
        3. Génère un script Playwright complet avec `import {{ test, expect }} from '@playwright/test';` et un bloc `test('{test_title}', async ({{ page }}) => {{ ... }});`.
        4. GESTION DES URLs : Si les données du test contiennent une URL absolue complète (commençant par http ou https, ex: `https://www.google.com/`), TU DOIS IMPÉRATIVEMENT utiliser cette URL complète dans `page.goto()`. N'utilise une URL relative (ex: `/login`) QUE si l'utilisateur n'a pas fourni de nom de domaine.
        5. Sois TRES ROBUSTE : Les sites publics ont souvent des popups "Accepter les cookies". Avant chaque première interaction importante, ajoute : `await page.locator('#L2AGLb, button:has-text("Tout accepter"), [id*="cookie"] button:has-text("accepter")').first().click({{ timeout: 5000 }}).catch(() => {{}});`.
        6. Déduis des sélecteurs très flexibles (ex: `page.locator('textarea[name="q"], input[name="q"], [title="Rechercher"]').first()` pour la barre de recherche Google). Préfère `locator` avec plusieurs sélecteurs CSS séparés par des virgules, ET AJOUTE TOUJOURS `.first()` à la fin de tes locators génériques. ATTENTION : N'utilise JAMAIS de sélecteur d'attribut trop généraliste sans préciser la balise pour des éléments cliquables ou interactifs (ex: n'utilise pas `[id*="login"]` ou `[class*="login"]` seul, car cela pourrait cibler un conteneur/wrapper parent comme un `div` ou un `form` au lieu du bouton lui-même. Utilise plutôt `button[id*="login"]` ou `button[type="submit"]`).
        7. Gère correctement les conditions ("si... apparait"). Exemple : si le texte dit "si la popup apparait", génère : `await page.waitForTimeout(2000); if (await page.locator('...').isVisible()) {{ await page.locator('...').click(); }}`.
        8. Inclus les assertions pour les 'Attendus'. Par exemple, `await expect(locator.first()).toBeVisible({{ timeout: 10000 }});` ou `await expect(page).not.toHaveURL(/.*login/, {{ timeout: 10000 }});`. Attention à la syntaxe stricte de Playwright (ex: `toHaveURL` prend une string/regex en premier argument).
        9. NE RENVOIE QUE LE CODE SOURCE. Pas d'explications avant ni après. Pas de blocs ```typescript``` ou ```javascript```.
        10. SOLUTION ANTI-TIMEOUT : Pour éviter que Playwright ne bloque sur des boutons masqués par du CSS ou des dropdowns, AJOUTE TOUJOURS `{{ force: true }}` lors de tes `click()` (ex: `await locator.click({{ force: true }})`). S'il s'agit d'une barre de recherche (comme Google), utilise TOUJOURS `await page.keyboard.press('Enter')` plutôt que de chercher un bouton "Recherche" !
        11. SAISIE DE TEXTE ROBUSTE : Pour tous les champs de saisie (input, textarea), utilise TOUJOURS la méthode native Playwright `await locator.first().fill('valeur')`. N'utilise JAMAIS `.evaluate()` pour remplir des champs — cette méthode ne fonctionne pas en mode headed et ne déclenche pas correctement les événements du navigateur. Après le fill(), si nécessaire, déclenche : `await locator.first().dispatchEvent('input')`.
        12. INTERACTION AVEC LES MENUS DÉROULANTS (SELECT) : Si une étape implique de choisir une option dans un menu déroulant, utilise TOUJOURS `await locator.first().selectOption({{ label: "Nom de l'option" }})` ou `selectOption({{ value: "Valeur" }})` au lieu d'un simple clic.
        13. CASES À COCHER & BOUTONS RADIO : Pour cocher une option ou case, utilise `await locator.first().check({{ force: true }})`.
        14. INTERACTION AVEC LES IFRAMES : Si les étapes indiquent d'agir dans un sous-cadre ou un iframe, utilise le sélecteur `page.frameLocator('iframe_selector').locator('element_selector')`.
        15. ATTENTE DE NAVIGATION : Après avoir cliqué sur un bouton de soumission ou une redirection de page, utilise `await page.waitForLoadState('networkidle');` pour s'assurer du chargement avant les assertions.
        16. ASSERTIONS PRÉCISES : Pour vérifier le CONTENU d'une page, n'utilise JAMAIS `page.locator('p, div, span').first()`. Utilise toujours un sélecteur sémantique stable (comme `main, article, [role="main"]` ou `body`).
        17. RESTRICTION DES TARGETS DE CLIC : Restreins toujours les cibles de clic aux éléments réellement interactifs (boutons `button`, liens `a`, ou inputs de type button/submit). Ne cible jamais de balises génériques de mise en page comme `div`, `form`, `section`, `span`, `p` ou `li` pour effectuer un clic, sauf si aucun autre élément n'est disponible et que l'étape le mentionne explicitement.
        18. IMAGES ET LOGOS : Pour vérifier qu'une image ou un logo est visible, utilise TOUJOURS son attribut `alt` directement : `page.locator('img[alt="Nom"]')`. N'utilise JAMAIS `.filter({{ hasText: ... }})` sur une balise `<img>` — les images n'ont pas de contenu texte et ce filtre retournera toujours 0 éléments. Exemple correct : `await expect(page.locator('img[alt="Google"]')).toBeVisible({{ timeout: 10000 }});`.
        19. FILTRE hasText : N'utilise `.filter({{ hasText: ... }})` QUE sur des éléments qui contiennent réellement du texte visible (boutons, liens, paragraphes, titres). Ne l'utilise JAMAIS sur `img`, `svg`, `video`, `input`, `canvas`, ou tout autre élément non-textuel.
        """
        
        messages = [{"role": "user", "content": prompt}]
        code = self._get_completion_with_fallback(messages, temperature=0.1)
        if code.startswith("```"):
            code = "\n".join(code.split("\n")[1:-1])

        # Post-processing automatique pour assainir les sélecteurs
        code = self._sanitize_playwright_selectors(code)
        return code

    def _sanitize_playwright_selectors(self, code):
        import re
        
        def split_selectors(selector_str):
            parts = []
            current = []
            depth = 0
            in_quotes = False
            quote_char = None
            
            for char in selector_str:
                if in_quotes:
                    if char == quote_char:
                        in_quotes = False
                    current.append(char)
                else:
                    if char in ['"', "'"]:
                        in_quotes = True
                        quote_char = char
                        current.append(char)
                    elif char in ['(', '[']:
                        depth += 1
                        current.append(char)
                    elif char in [')', ']']:
                        depth -= 1
                        current.append(char)
                    elif char == ',' and depth == 0:
                        parts.append("".join(current).strip())
                        current = []
                    else:
                        current.append(char)
            if current:
                parts.append("".join(current).strip())
            return parts

        def sanitize_part(part):
            stripped = part.strip()
            if not stripped:
                return part
                
            # Trouver le dernier combinateur à la racine
            combinators = [' ', '>', '+', '~']
            last_combinator_idx = -1
            depth = 0
            in_quotes = False
            quote_char = None
            
            for i, char in enumerate(stripped):
                if in_quotes:
                    if char == quote_char:
                        in_quotes = False
                else:
                    if char in ['"', "'"]:
                        in_quotes = True
                        quote_char = char
                    elif char in ['(', '[']:
                        depth += 1
                    elif char in [')', ']']:
                        depth -= 1
                    elif char in combinators and depth == 0:
                        last_combinator_idx = i
                        
            if last_combinator_idx != -1:
                prefix = stripped[:last_combinator_idx + 1]
                last_token = stripped[last_combinator_idx + 1:].strip()
            else:
                prefix = ""
                last_token = stripped
                
            if last_token.startswith('[') or last_token.startswith('.') or last_token.startswith('#'):
                lower_token = last_token.lower()
                if 'role=' in lower_token or 'type=' in lower_token:
                    return part
                
                # 1. Champs de saisie (Inputs)
                input_keywords = ['username', 'password', 'email', 'text', 'input', 'valeur', 'search', 'query', 'q', 'pwd', 'usr', 'champ', 'field', 'recherche']
                is_input_context = any(kw in lower_token for kw in input_keywords)
                
                if is_input_context:
                    expanded = []
                    for tag in ['input', 'textarea', 'select']:
                        expanded.append(f"{prefix}{tag}{last_token}")
                    return ", ".join(expanded)
                    
                # 2. Titres/Textes
                heading_keywords = ['heading', 'title', 'header', 'titre', 'text', 'label', 'message', 'alert', 'success', 'error', 'fail', 'status', 'welcome', 'secure', 'area', 'info', 'banner']
                is_heading_context = any(kw in lower_token for kw in heading_keywords)
                
                if is_heading_context:
                    expanded = []
                    for tag in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6', '[role="heading"]', 'span', 'p']:
                        expanded.append(f"{prefix}{tag}{last_token}")
                    return ", ".join(expanded)
                    
                # 3. Actions de Clic
                expanded = []
                for tag in ['button', 'a', 'input', '[role="button"]', '[role="link"]']:
                    expanded.append(f"{prefix}{tag}{last_token}")
                return ", ".join(expanded)
                
            return part

        def replace_match(match):
            quote = match.group(1)
            selector_str = match.group(2)
            
            parts = split_selectors(selector_str)
            sanitized_parts = []
            for part in parts:
                sanitized_parts.append(sanitize_part(part))
                
            new_selector_str = ", ".join(sanitized_parts)
            new_selector_str = re.sub(r'\s*,\s*', ', ', new_selector_str)
            return f"page.locator({quote}{new_selector_str}{quote})"

        pattern = r'page\.locator\(\s*(["\'])(.*?)\1\s*\)'
        return re.sub(pattern, replace_match, code)

    def generate_anomaly_from_logs(self, test_title, logs):
        """
        Génère un titre et une description d'anomalie pertinents à partir des logs d'erreur Playwright.
        Distingue un bug de script (SCRIPT_ERROR) d'un vrai bug applicatif (APP_BUG).
        """
        prompt = f"""
        Tu es un expert QA senior. Un test Playwright automatisé vient d'échouer.
        Titre du Cas de Test : {test_title}
        
        LOGS D'ERREUR PLAYWRIGHT :
        {logs[-2000:]}
        
        TON TRAVAIL EN 2 ÉTAPES :
        
        ÉTAPE 1 — Diagnostique la CAUSE RACINE avec précision parmi ces catégories :
        
        1. SCRIPT_ERROR (Erreur liée au code du script de test) :
           - Sélecteurs introuvables / Erreurs de syntaxe (ex: "Unable to find element", "Target closed", "locator.click: Timeout...").
           - Mauvais ciblage : Par exemple, cliquer sur un conteneur parent (div/form) au lieu du bouton interne.
           - Sélecteurs obsolètes : L'élément dans le script n'existe pas ou a été renommé, sans que l'application ne présente d'erreur.
           
        2. APP_BUG (Erreur de l'application ou Échec Fonctionnel) :
           - ERREURS DE REDIRECTION / AUTHENTIFICATION : Si le test s'attend à une page/élément cible (ex: attend "Secure Area" ou "Dashboard") mais que la page n'a pas changé (reçoit toujours "Login Page", "Connexion" ou reste sur l'URL d'origine), c'est un échec d'action (identifiants incorrects ou rejetés). Titrez : "Échec de connexion" ou "Échec d'authentification".
           - ERREURS HTTP / SERVEUR : Si les logs mentionnent des codes comme 500, 502, 503, 404 ou des crashs de console. Titrez : "Erreur Serveur HTTP [Code]" ou "Page introuvable (404)".
           - ERREURS DE VALIDATION : Si le test reste bloqué sur un formulaire avec des messages d'erreur textuels de validation (ex: "champ obligatoire", "format invalide", "déjà existant"). Titrez : "Erreur de validation de formulaire".
           - TIMEOUT SUR ÉLÉMENT FONCTIONNEL : Si la page a mis trop de temps à charger un élément métier essentiel (hors problème de sélecteur technique). Titrez : "Timeout de chargement applicatif".
        
        ÉTAPE 2 — Retourne UNIQUEMENT ce JSON strict :
        {{"cause": "SCRIPT_ERROR ou APP_BUG", "titre": "Un titre d'anomalie court et extrêmement explicite", "description": "Une description claire en Français de ce qui s'est réellement passé. Si SCRIPT_ERROR : expliquez l'erreur de code de test. Si APP_BUG : expliquez le comportement ou le rejet de l'application (ex: échec d'authentification ou erreur 500)."}}
        Ne renvoie rien d'autre que le JSON valide.
        """
        try:
            messages = [{"role": "user", "content": prompt}]
            response = self._get_completion_with_fallback(messages, temperature=0.1)
            if response.startswith("```json"):
                response = response.replace("```json", "").replace("```", "").strip()
            if response.startswith("```"):
                response = response.replace("```", "").strip()
                
            import json
            data = json.loads(response)
            cause = data.get("cause", "APP_BUG")
            titre = data.get("titre", f"Échec automatique : {test_title}")
            description = data.get("description", logs[-1000:])
            
            return titre, description
        except Exception as e:
            return f"Échec d'exécution : {test_title}", f"Le test automatisé a échoué. \n\nLogs d'erreur:\n{logs[-1000:]}"

